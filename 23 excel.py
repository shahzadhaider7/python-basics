import openpyxl

workbook = openpyxl.load_workbook("E:\Employees.xlsx")

workbook.properties     # displays the properties of this excel file

workbook.sheetnames     # displays the names of sheets in this excel file

workbook.active     # displays the active sheet in the excel file (usually first sheet)

sheet = workbook["EmployeeData"]

workbook.create_sheet("Test")   # for creating a sheet in the excel file

workbook.save("E:\Employees.xlsx")   # for saving the excel file after we created a new sheet

sheet = workbook["Test"]    # assigning the newly created sheet to a variable

workbook.remove(sheet)      # removing the newly created sheet

workbook.save("E:\Employees.xlsx")  # saving the excel file to save changes

del workbook["Test1"]       # deleting another newly created sheet by another method

workbook.save("E:\Employees.xlsx")   # saving excel file to save changes

sheet.active_cell       # sheet has the first sheet stored on it, which is "EmployeeData"

sheet.dimensions        # starting and ending box - (first and last, from A1 to F11)


# other sheet properties

sheet.sheet_format

sheet.sheet_properties

sheet.sheet_state       # state - visible or not in a workbook

sheet.sheet_view

dir(sheet)          # to view all available methods and attributes for sheet

sheet.max_row       # number of rows

sheet.max_column    # number of columns

for i in sheet.values:      # to print all the values in sheet and display them in tuples
    print(i)   


# for getting a value in a cell

sheet["B6"].value       # gets the value from column B and row 6

sheet.cell( row = 3 , column = 4 ).value    # another way of getting value from a cell

# how to store the value in a cell to a variable

my_cell = sheet["C4"]       # store the value in the C4 cell to my_cell

my_cell.value   # to access the value

my_cell.row     # it is in row no. 4

my_cell.column  # it is in column no. 3, which is C

my_cell.coordinate      # to get the full coordinate of the cell

my_cell.data_type       # to get the data type of this cell, i.e. "s" for string

my_cell.encoding         # to get the encoding type of cell, default = utf-8

my_cell1 = sheet["B2"]   # to assign the value in Cell B2 to my_cell1 object

my_cell1.value = "Shahzad"    # to change the value in the cell

workbook.save("E:\Employees.xlsx")  # saving changes

my_cell1.parent     # to know which sheet it belongs to


#  Cell styles using python openpyxl module


dir(openpyxl.styles)    # to list all available style attributes

from openpyxl.styles import *    # import all style attributes 

my_font = Font( color = colors.RED , bold = True , italic = True )   # setting font styles to a variable

new_cell = sheet["B8"]

new_cell.font = my_font   # applying the font styling variable to our cell


my_fill = PatternFill( fill_type = 'solid' , bgColor = "F6A3BC" )    # setting background styles to a variable

new_cell.fill = my_fill     # applying background styling variable to our cell

my_border = Border( left = Side( border_style = 'double' , color = '322FEC' ) , right = Side( border_style = 'double' , color = '322FEC' ) , top = Side( border_style = 'double' , color = '322FEC' ) , bottom = Side( border_style = 'double' , color = '322FEC' ) )    # setting styles of left border, right border, top border, bottom border and assigning to a variable

new_cell.border = my_border    # applying border styling variable to our cell

my_align = Alignment( horizontal = 'left' )     # setting alignment style to a variable

new_cell.alignment = my_align        # applying alignment styling variable to our cell


workbook.save("E:\Employees.xlsx")  # saving changes




